from __future__ import annotations

from datetime import date, datetime
from functools import lru_cache
from io import BytesIO
import json
from pathlib import Path
import re
import unicodedata
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


BASE_DIR = Path(__file__).resolve().parent
PROFILES_PATH = BASE_DIR / "format_profiles.json"


def list_sheet_names(file_bytes: bytes) -> list[str]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    return wb.sheetnames


def list_format_profiles() -> list[dict[str, str]]:
    config = load_profiles_config()
    output: list[dict[str, str]] = []
    for profile in config["profiles"]:
        output.append(
            {
                "id": profile["id"],
                "name": profile.get("name", profile["id"]),
                "description": profile.get("description", ""),
            }
        )
    return output


def detect_profile_id(file_bytes: bytes) -> str:
    config = load_profiles_config()
    profiles = config["profiles"]
    default_id = config["default_profile"]

    if len(profiles) == 1:
        return profiles[0]["id"]

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    text_pool = collect_normalized_text_pool(wb)

    best_profile_id = default_id
    best_score = -1

    for profile in profiles:
        tokens = [normalize_text(t) for t in profile.get("match_tokens", []) if t]
        if not tokens:
            continue
        score = sum(1 for token in tokens if token in text_pool)
        if score > best_score:
            best_profile_id = profile["id"]
            best_score = score

    return best_profile_id


def get_profile_field_order(profile_id: str | None = None) -> list[str]:
    profile = resolve_profile(profile_id)
    return list(profile["field_order"])


def list_data_sheet_names(file_bytes: bytes, profile_id: str | None = None) -> list[str]:
    profile = resolve_profile(profile_id)
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    names: list[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        row_map = build_row_map(ws, profile)
        sample_row_key = profile["sample_config"]["row_key"]
        sample_row = row_map.get(sample_row_key)
        if sample_row is None:
            continue
        sample_columns = detect_sample_columns(ws, sample_row, profile)
        if not sample_columns:
            continue
        if not has_data_in_any_sample(ws, row_map, sample_columns, profile):
            continue
        names.append(name)
    return names


def infer_sheet_headers(file_bytes: bytes, sheet_name: str, profile_id: str | None = None) -> list[str]:
    profile = resolve_profile(profile_id)
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo.")

    ws = wb[sheet_name]
    row_map = build_row_map(ws, profile)
    return build_headers_from_sheet(ws, row_map, profile)


def extract_sheet(
    file_bytes: bytes,
    sheet_name: str,
    tipo_producto: str = "no especificar",
    drop_empty_samples: bool = True,
    profile_id: str | None = None,
) -> list[dict[str, Any]]:
    profile = resolve_profile(profile_id)

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo.")

    ws = wb[sheet_name]
    row_map = build_row_map(ws, profile)

    sample_row_key = profile["sample_config"]["row_key"]
    sample_row = row_map.get(sample_row_key)
    if sample_row is None:
        raise ValueError("No se encontro la fila de muestras en la hoja seleccionada.")

    sample_columns = detect_sample_columns(ws, sample_row, profile)
    if not sample_columns:
        raise ValueError("No se encontraron columnas de muestra (N°1, N°2, etc.).")

    metadata = read_metadata(ws, profile)
    conv_value, org_value = product_type_marks(tipo_producto)

    extracted: list[dict[str, Any]] = []

    for col_idx in sample_columns:
        record = {
            "fecha_analisis": metadata.get("fecha_analisis"),
            "cliente": metadata.get("cliente"),
            "turno": metadata.get("turno"),
            "fecha_produccion": metadata.get("fecha_produccion"),
            "formato": metadata.get("formato"),
            "numero_hoja": metadata.get("numero_hoja"),
            "convencional": conv_value,
            "organico": org_value,
            "variedad": metadata.get("variedad"),
            "tamano_muestra": metadata.get("tamano_muestra"),
            "muestras": cell_value(ws, row_map.get("muestras"), col_idx),
            "hora": cell_value(ws, row_map.get("hora"), col_idx),
            "lote_mp": cell_value(ws, row_map.get("lote_mp"), col_idx),
            "codigo_produccion": cell_value(ws, row_map.get("codigo_produccion"), col_idx),
            "pallet_pt": cell_value(ws, row_map.get("pallet_pt"), col_idx),
            "ph": cell_value(ws, row_map.get("ph"), col_idx),
            "bx": cell_value(ws, row_map.get("bx"), col_idx),
            "color_decolor": cell_value(ws, row_map.get("color_decolor"), col_idx),
            "color_palido": cell_value(ws, row_map.get("color_palido"), col_idx),
            "tamano_irregular": cell_value(ws, row_map.get("tamano_irregular"), col_idx),
            "tamano_menor": cell_value(ws, row_map.get("tamano_menor"), col_idx),
            "tamano_mayor": cell_value(ws, row_map.get("tamano_mayor"), col_idx),
            "tamano_sumatoria": cell_value(ws, row_map.get("tamano_sumatoria"), col_idx),
            "defecto_sobremaduro": cell_value(ws, row_map.get("defecto_sobremaduro"), col_idx),
            "defecto_decoloracion": cell_value(ws, row_map.get("defecto_decoloracion"), col_idx),
            "defecto_danio": cell_value(ws, row_map.get("defecto_danio"), col_idx),
            "defecto_aglomerado": cell_value(ws, row_map.get("defecto_aglomerado"), col_idx),
            "defecto_piel": cell_value(ws, row_map.get("defecto_piel"), col_idx),
            "defecto_sumatoria": cell_value(ws, row_map.get("defecto_sumatoria"), col_idx),
            "serio_piel": cell_value(ws, row_map.get("serio_piel"), col_idx),
            "serio_semilla": cell_value(ws, row_map.get("serio_semilla"), col_idx),
            "serio_mat_vegetal": cell_value(ws, row_map.get("serio_mat_vegetal"), col_idx),
            "serio_mat_extrana": cell_value(ws, row_map.get("serio_mat_extrana"), col_idx),
            "motoso": cell_value(ws, row_map.get("motoso"), col_idx),
            "duro": cell_value(ws, row_map.get("duro"), col_idx),
            "otros_1": cell_value(ws, row_map.get("otros"), col_idx),
            "otros_2": cell_value(ws, safe_row(row_map.get("otros"), 1), col_idx),
            "otros_3": cell_value(ws, safe_row(row_map.get("otros"), 2), col_idx),
            "otros_4": cell_value(ws, safe_row(row_map.get("otros"), 3), col_idx),
            "temperatura": cell_value(ws, row_map.get("temperatura"), col_idx),
            "sabor": cell_value(ws, row_map.get("sabor"), col_idx),
        }

        if drop_empty_samples and is_empty_sample(record, profile):
            continue

        extracted.append(record)

    return extracted


def extract_sheets(
    file_bytes: bytes,
    sheet_names: list[str],
    tipo_producto: str = "no especificar",
    drop_empty_samples: bool = True,
    profile_id: str | None = None,
) -> list[dict[str, Any]]:
    all_rows: list[dict[str, Any]] = []
    for sheet_name in sheet_names:
        rows = extract_sheet(
            file_bytes=file_bytes,
            sheet_name=sheet_name,
            tipo_producto=tipo_producto,
            drop_empty_samples=drop_empty_samples,
            profile_id=profile_id,
        )
        all_rows.extend(rows)
    return all_rows


def render_output_excel(
    records: list[dict[str, Any]],
    output_headers: list[str],
    field_order: list[str],
    template_bytes: bytes | None = None,
) -> bytes:
    if template_bytes:
        wb = load_workbook(BytesIO(template_bytes))
        ws = wb[wb.sheetnames[0]]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"

    write_headers(ws, output_headers)
    clear_data_rows(ws)

    effective_order = field_order[: len(output_headers)]
    for record in records:
        ws.append([record.get(field_id) for field_id in effective_order])

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def rows_to_display_records(
    records: list[dict[str, Any]],
    output_headers: list[str],
    field_order: list[str],
) -> list[dict[str, Any]]:
    effective_order = field_order[: len(output_headers)]
    output: list[dict[str, Any]] = []
    for row in records:
        output.append({header: row.get(field_id) for field_id, header in zip(effective_order, output_headers, strict=False)})
    return output


def normalize_records_for_output(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in records:
        clean_row: dict[str, Any] = {}
        for key, value in row.items():
            if is_effectively_empty(value):
                clean_row[key] = None
            else:
                clean_row[key] = value
        normalized.append(clean_row)
    return normalized


def write_headers(ws: Worksheet, output_headers: list[str]) -> None:
    for idx, header in enumerate(output_headers, start=1):
        ws.cell(1, idx).value = header

    if ws.max_column > len(output_headers):
        for idx in range(len(output_headers) + 1, ws.max_column + 1):
            ws.cell(1, idx).value = None


def clear_data_rows(ws: Worksheet) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def build_headers_from_sheet(ws: Worksheet, row_map: dict[str, int | None], profile: dict[str, Any]) -> list[str]:
    header_cells = profile.get("header_cells", {})

    base_fqo = row_label(ws, row_map.get("ph"), 1, "FQO")
    base_color = row_label(ws, row_map.get("color_decolor"), 1, "COLOR")
    base_tamano = row_label(ws, row_map.get("tamano_irregular"), 1, "TAMAÑO")

    defect_word = row_label(ws, row_map.get("defecto_sobremaduro"), 1, "DEFECTOS")
    minor_word = row_label(ws, row_map.get("defecto_sobremaduro"), 2, "MENORES")
    serious_word = row_label(ws, row_map.get("serio_piel"), 2, "SERIOS")

    minor_prefix = join_header([defect_word, minor_word], "DEFECTOS MENORES")
    serious_prefix = join_header([defect_word, serious_word], "DEFECTOS SERIOS")

    headers = [
        label_by_ref(ws, header_cells.get("fecha_analisis"), "FECHA DE ANALISIS"),
        label_by_ref(ws, header_cells.get("cliente"), "CLIENTE"),
        label_by_ref(ws, header_cells.get("turno"), "TURNO"),
        label_by_ref(ws, header_cells.get("fecha_produccion"), "FECHA DE PRODUCCION"),
        label_by_ref(ws, header_cells.get("formato"), "FORMATO"),
        label_by_ref(ws, header_cells.get("numero_hoja"), "N° HOJA"),
        label_by_ref(ws, header_cells.get("convencional"), "CONVENCIONAL"),
        label_by_ref(ws, header_cells.get("organico"), "ORGANICO"),
        label_by_ref(ws, header_cells.get("variedad"), "VARIEDAD"),
        label_by_ref(ws, header_cells.get("tamano_muestra"), "TAMAÑO DE MUESTRA"),
        row_label(ws, row_map.get("muestras"), 1, "MUESTRAS"),
        row_label(ws, row_map.get("hora"), 1, "HORA"),
        row_label(ws, row_map.get("lote_mp"), 1, "LOTE M.P."),
        row_label(ws, row_map.get("codigo_produccion"), 1, "CODIGO PRODUCCION"),
        row_label(ws, row_map.get("pallet_pt"), 1, "N° PALLET PT"),
        join_header([base_fqo, row_label(ws, row_map.get("ph"), 2, "PH"), label_at_row(ws, row_map.get("ph"), 5)], "FQO PH"),
        join_header([base_fqo, row_label(ws, row_map.get("bx"), 2, "°BX"), label_at_row(ws, row_map.get("bx"), 5)], "FQO °BX"),
        join_header([base_color, row_label(ws, row_map.get("color_decolor"), 2, "DECOLOR")], "COLOR DECOLOR"),
        join_header([base_color, row_label(ws, row_map.get("color_palido"), 2, "PALIDO")], "COLOR PALIDO"),
        join_header([base_tamano, row_label(ws, row_map.get("tamano_irregular"), 2, "IRREGULAR")], "TAMAÑO IRREGULAR"),
        join_header([base_tamano, row_label(ws, row_map.get("tamano_menor"), 2, "< MM 20 X 20")], "TAMAÑO < MM 20 X 20"),
        join_header([base_tamano, row_label(ws, row_map.get("tamano_mayor"), 2, "30 X 30 > MM")], "TAMAÑO 30 X 30 > MM"),
        join_header([base_tamano, row_label(ws, row_map.get("tamano_sumatoria"), 2, "SUMATORIA")], "TAMAÑO SUMATORIA"),
        join_header([minor_prefix, row_label(ws, row_map.get("defecto_sobremaduro"), 3, "SOBREMADURO")], "DEFECTOS MENORES SOBREMADURO"),
        join_header([minor_prefix, row_label(ws, row_map.get("defecto_decoloracion"), 3, "DECOLORACION")], "DEFECTOS MENORES DECOLORACION"),
        join_header([minor_prefix, row_label(ws, row_map.get("defecto_danio"), 3, "DAÑO")], "DEFECTOS MENORES DAÑO"),
        join_header([minor_prefix, row_label(ws, row_map.get("defecto_aglomerado"), 3, "AGLOMERADO")], "DEFECTOS MENORES AGLOMERADO"),
        join_header([minor_prefix, row_label(ws, row_map.get("defecto_piel"), 3, "PIEL")], "DEFECTOS MENORES PIEL"),
        join_header([minor_prefix, row_label(ws, row_map.get("defecto_sumatoria"), 3, "SUMATORIA")], "DEFECTOS MENORES SUMATORIA"),
        join_header([serious_prefix, row_label(ws, row_map.get("serio_piel"), 3, "PIEL")], "DEFECTOS SERIOS PIEL"),
        join_header([serious_prefix, row_label(ws, row_map.get("serio_semilla"), 3, "SEMILLA")], "DEFECTOS SERIOS SEMILLA"),
        join_header([serious_prefix, row_label(ws, row_map.get("serio_mat_vegetal"), 3, "MAT. VEGETAL")], "DEFECTOS SERIOS MAT. VEGETAL"),
        join_header([serious_prefix, row_label(ws, row_map.get("serio_mat_extrana"), 3, "MAT. EXTRAÑA")], "DEFECTOS SERIOS MAT. EXTRAÑA"),
        join_header([defect_word, row_label(ws, row_map.get("motoso"), 2, "MOTOSO")], "DEFECTOS MOTOSO"),
        join_header([defect_word, row_label(ws, row_map.get("duro"), 2, "DURO")], "DEFECTOS DURO"),
        join_header([defect_word, row_label(ws, row_map.get("otros"), 2, "OTROS"), "1"], "DEFECTOS OTROS 1"),
        join_header([defect_word, row_label(ws, safe_row(row_map.get("otros"), 1), 2, "OTROS"), "2"], "DEFECTOS OTROS 2"),
        join_header([defect_word, row_label(ws, safe_row(row_map.get("otros"), 2), 2, "OTROS"), "3"], "DEFECTOS OTROS 3"),
        join_header([defect_word, row_label(ws, safe_row(row_map.get("otros"), 3), 2, "OTROS"), "4"], "DEFECTOS OTROS 4"),
        row_label(ws, row_map.get("temperatura"), 1, "TEMPERATURA"),
        row_label(ws, row_map.get("sabor"), 1, "SABOR Y OLOR"),
    ]

    headers = unique_headers(headers)
    field_count = len(profile["field_order"])
    if len(headers) > field_count:
        return headers[:field_count]
    if len(headers) < field_count:
        padded = headers[:]
        while len(padded) < field_count:
            padded.append(f"COLUMNA_{len(padded) + 1}")
        return padded
    return headers


def read_metadata(ws: Worksheet, profile: dict[str, Any]) -> dict[str, Any]:
    metadata: dict[str, Any] = {}
    metadata_cells = profile.get("metadata_cells", {})
    date_fields = set(profile.get("metadata_date_fields", []))

    for key, cell_ref in metadata_cells.items():
        value = cell_value_by_ref(ws, cell_ref)
        if key in date_fields:
            metadata[key] = parse_excel_date(value)
        else:
            metadata[key] = clean_value(value)

    return metadata


def product_type_marks(tipo_producto: str) -> tuple[str | None, str | None]:
    key = normalize_text(tipo_producto)
    if key.startswith("CONV"):
        return "X", None
    if key.startswith("ORG"):
        return None, "X"
    return None, None


def cell_value(ws: Worksheet, row: int | None, col: int) -> Any:
    if row is None:
        return None
    return clean_value(ws.cell(row, col).value)


def cell_value_by_ref(ws: Worksheet, cell_ref: str | None) -> Any:
    if not cell_ref:
        return None
    return ws[cell_ref].value


def is_empty_sample(record: dict[str, Any], profile: dict[str, Any]) -> bool:
    signal_fields = profile.get("signal_row_keys", [])
    for field in signal_fields:
        if not is_effectively_empty(record.get(field)):
            return False
    return True


def has_data_in_any_sample(
    ws: Worksheet,
    row_map: dict[str, int | None],
    sample_columns: list[int],
    profile: dict[str, Any],
) -> bool:
    signal_rows = [row_map.get(key) for key in profile.get("signal_row_keys", [])]
    for col_idx in sample_columns:
        for row_idx in signal_rows:
            if row_idx is None:
                continue
            value = clean_value(ws.cell(row_idx, col_idx).value)
            if not is_effectively_empty(value):
                return True
    return False


def build_row_map(ws: Worksheet, profile: dict[str, Any]) -> dict[str, int | None]:
    rules = profile.get("row_rules", {})
    rows: dict[str, int | None] = {}

    for key, rule in rules.items():
        found_row: int | None = None

        find_rule = rule.get("find")
        if find_rule:
            found_row = resolve_find_rule(ws, rows, find_rule)

        offset_from = rule.get("offset_from")
        if found_row is None and offset_from:
            base_row = rows.get(offset_from)
            if base_row is not None:
                candidate = base_row + int(rule.get("offset", 0))
                offset_tokens = rule.get("offset_tokens")
                offset_columns = tuple(rule.get("offset_columns", [1, 2, 3]))
                if offset_tokens:
                    if row_has_tokens(ws, candidate, offset_tokens, offset_columns):
                        found_row = candidate
                else:
                    found_row = candidate

        fallback_find = rule.get("fallback_find")
        if found_row is None and fallback_find:
            found_row = resolve_find_rule(ws, rows, fallback_find)

        fallback_from = rule.get("fallback_from")
        if found_row is None and fallback_from:
            base_row = rows.get(fallback_from)
            if base_row is not None:
                found_row = base_row + int(rule.get("fallback_offset", 0))

        rows[key] = found_row

    return rows


def resolve_find_rule(ws: Worksheet, rows: dict[str, int | None], find_rule: dict[str, Any]) -> int | None:
    tokens = find_rule.get("tokens", [])
    columns = tuple(find_rule.get("columns", [1, 2, 3]))
    min_row = int(find_rule.get("min_row", 1))
    max_row = int(find_rule.get("max_row", 90))

    min_row_from = find_rule.get("min_row_from")
    if min_row_from:
        ref_row = rows.get(min_row_from)
        if ref_row is not None:
            min_row = max(min_row, ref_row)

    return find_row(ws, tokens=tokens, columns=columns, min_row=min_row, max_row=max_row)


def row_has_tokens(
    ws: Worksheet,
    row: int,
    tokens: Iterable[str],
    columns: tuple[int, ...],
) -> bool:
    lowered_tokens = [normalize_text(t) for t in tokens]
    for col in columns:
        value = normalize_text(ws.cell(row, col).value)
        if value and all(token in value for token in lowered_tokens):
            return True
    return False


def find_row(
    ws: Worksheet,
    tokens: list[str],
    columns: tuple[int, ...] = (1, 2, 3),
    min_row: int = 1,
    max_row: int = 90,
) -> int | None:
    lowered_tokens = [normalize_text(token) for token in tokens]

    for row in range(min_row, max_row + 1):
        for col in columns:
            value = normalize_text(ws.cell(row, col).value)
            if value and all(token in value for token in lowered_tokens):
                return row

    return None


def detect_sample_columns(ws: Worksheet, sample_row: int, profile: dict[str, Any]) -> list[int]:
    sample_cfg = profile.get("sample_config", {})
    start_col = int(sample_cfg.get("start_col", 6))
    pattern = sample_cfg.get("header_regex", r"N\s*°?\s*\d+")

    cols: list[int] = []
    for col in range(start_col, ws.max_column + 1):
        header = clean_value(ws.cell(sample_row, col).value)
        if header is None:
            continue
        text = str(header).upper()
        if re.search(pattern, text):
            cols.append(col)

    if cols:
        return cols

    for col in range(start_col, ws.max_column + 1):
        header = clean_value(ws.cell(sample_row, col).value)
        if header is not None:
            cols.append(col)
    return cols


def parse_excel_date(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value

    text = clean_value(value)
    if not text:
        return None

    match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", str(text))
    if not match:
        return text

    day, month, year = map(int, match.groups())
    try:
        return datetime(year, month, day)
    except ValueError:
        return text


def clean_value(value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, str):
        cleaned = " ".join(value.split()).strip()
        if not cleaned:
            return None
        return cleaned

    return value


def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return text.strip()


def collect_normalized_text_pool(wb) -> str:
    parts: list[str] = []
    for ws in wb.worksheets:
        max_row = min(ws.max_row, 60)
        max_col = min(ws.max_column, 20)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                value = ws.cell(row, col).value
                if value is None:
                    continue
                parts.append(normalize_text(value))
    return " ".join(parts)


def safe_row(base_row: int | None, offset: int) -> int | None:
    if base_row is None:
        return None
    return base_row + offset


def is_effectively_empty(value: Any) -> bool:
    if value is None:
        return True
    if not isinstance(value, str):
        return False

    cleaned = " ".join(value.split()).strip()
    if not cleaned:
        return True
    if re.fullmatch(r"[-.`ˋ]+", cleaned):
        return True
    if normalize_text(cleaned) in {"NA", "N A", "NONE", "NULL"}:
        return True
    return False


def label_by_ref(ws: Worksheet, cell_ref: str | None, fallback: str) -> str:
    if cell_ref:
        return normalize_header_token(ws[cell_ref].value, fallback)
    return fallback


def label_at_row(ws: Worksheet, row: int | None, col: int) -> str | None:
    if row is None:
        return None
    return normalize_header_token(ws.cell(row, col).value)


def row_label(ws: Worksheet, row: int | None, col: int, fallback: str) -> str:
    if row is None:
        return fallback
    return normalize_header_token(ws.cell(row, col).value, fallback)


def normalize_header_token(value: Any, fallback: str | None = None) -> str:
    if value is None:
        return fallback or ""

    text = str(value)
    if "\n" in text:
        text = text.splitlines()[0]

    text = " ".join(text.split()).strip().rstrip(":")
    if not text:
        return fallback or ""

    return text.upper()


def join_header(parts: Iterable[Any], fallback: str) -> str:
    tokens: list[str] = []
    for part in parts:
        token = normalize_header_token(part)
        if token:
            tokens.append(token)

    if not tokens:
        return fallback

    return " ".join(tokens)


def unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []

    for raw_header in headers:
        header = raw_header or "COLUMNA"
        count = seen.get(header, 0) + 1
        seen[header] = count
        if count == 1:
            result.append(header)
        else:
            result.append(f"{header}_{count}")

    return result


@lru_cache(maxsize=1)
def load_profiles_config() -> dict[str, Any]:
    if not PROFILES_PATH.exists():
        raise FileNotFoundError(f"No se encontró el archivo de perfiles: {PROFILES_PATH}")

    data = json.loads(PROFILES_PATH.read_text(encoding="utf-8-sig"))
    validate_profiles_config(data)
    return data


def resolve_profile(profile_id: str | None = None) -> dict[str, Any]:
    config = load_profiles_config()
    target = profile_id or config["default_profile"]

    for profile in config["profiles"]:
        if profile["id"] == target:
            return profile

    raise ValueError(f"Perfil no encontrado: {target}")


def validate_profiles_config(data: dict[str, Any]) -> None:
    if "default_profile" not in data or "profiles" not in data:
        raise ValueError("format_profiles.json debe incluir 'default_profile' y 'profiles'.")

    if not isinstance(data["profiles"], list) or not data["profiles"]:
        raise ValueError("'profiles' debe ser una lista no vacía.")

    ids = {profile.get("id") for profile in data["profiles"]}
    if data["default_profile"] not in ids:
        raise ValueError("'default_profile' no coincide con ningún perfil definido.")

    required_profile_keys = [
        "id",
        "field_order",
        "metadata_cells",
        "sample_config",
        "row_rules",
    ]

    for profile in data["profiles"]:
        for key in required_profile_keys:
            if key not in profile:
                raise ValueError(f"El perfil '{profile.get('id', '?')}' no tiene la clave requerida '{key}'.")
